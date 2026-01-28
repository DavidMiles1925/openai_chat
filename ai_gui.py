from openai import OpenAI
import threading
import tkinter as tk
from tkinter import scrolledtext, filedialog, Menu, messagebox
from datetime import datetime
import os
from docx import Document
from pypdf import PdfReader
import base64
import urllib.request
import traceback
import tkinter.font as tkfont
from io import BytesIO

# Pillow for image loading & conversion to Tkinter PhotoImage
from PIL import Image, ImageTk

# keyring (use OS-provided secure storage). If missing, user must pip install keyring.
try:
    import keyring
except Exception as e:
    keyring = None
    print("Warning: keyring module not available. Install with: pip install keyring")

from config import DEFAULT_MODEL_VERSION, ASSISTANT_NAME, IMAGE_MODEL_VERSION, AVAILABLE_MODELS, MAX_UPLOADED_IMAGE_DIMENSION, VISION_CAPABLE_MODELS

try:
    RESAMPLE_FILTER = Image.Resampling.LANCZOS
except AttributeError:
    RESAMPLE_FILTER = getattr(Image, "LANCZOS", Image.BICUBIC)

# -----------------------------
# API key / keyring management
# -----------------------------
SERVICE_NAME = "gpt_gui_de"          # change if you want a different keychain service name
KEYRING_USERNAME = "openai_api_key"

# Track the session key (may be session-only); load_api_key may not see session-only keys,
# so this tracks whether the app currently has a usable key in-memory for the session.
current_session_key = None

def load_api_key():
    """
    Try keyring first, then fall back to the OPENAI_KEY environment variable.
    Returns the key string or None.
    """
    # Try keyring if available
    if keyring:
        try:
            k = keyring.get_password(SERVICE_NAME, KEYRING_USERNAME)
            if k:
                return k
        except Exception:
            # ignore keyring issues and fallback to environment
            pass

    # Fallback to env var
    env_key = os.getenv("OPENAI_KEY")
    if env_key and env_key != "No_Key":
        return env_key
    return None

def save_api_key_in_keyring(key):
    if not keyring:
        print("Keyring not available: cannot save key.")
        return False
    try:
        keyring.set_password(SERVICE_NAME, KEYRING_USERNAME, key)
        return True
    except Exception as e:
        print("Keyring save error:", e)
        return False

def delete_api_key_from_keyring():
    if not keyring:
        return
    try:
        keyring.delete_password(SERVICE_NAME, KEYRING_USERNAME)
    except Exception:
        pass

def init_client_with_key(key):
    """
    Re-initialize the global OpenAI client using key (or none).
    Also updates current_session_key.
    """
    global client, current_session_key
    try:
        if key:
            client = OpenAI(api_key=key)
            current_session_key = key
        else:
            # No explicit key -> let SDK use env var if present
            client = OpenAI()
            # set current_session_key to None (session doesn't have an explicit key)
            current_session_key = None
    except Exception as e:
        print("Failed to initialize OpenAI client:", e)
        # Keep previous client if any; do not alter current_session_key on failure

# -----------------------------
# Validation modal & helpers
# -----------------------------
def _validate_key_worker(key, remember_flag, dlg):
    """
    Validate the provided key by performing a tiny API call.
    Runs in a background thread. Uses app.after to update the UI.
    """
    err = None
    err_msg = None
    try:
        temp_client = OpenAI(api_key=key)
        # Small test call (very small) — use new param name max_completion_tokens
        temp_client.chat.completions.create(
            model=_default_model,
            messages=[{"role": "user", "content": "Say ok in one word"}],
            max_completion_tokens=10
        )
    except Exception as e:
        err = e
        err_msg = str(e)

    def on_fail():
        try:
            dlg.status_var.set("Validation failed.")
            # Re-enable dialog controls
            dlg.save_btn.config(state='normal')
            dlg.cancel_btn.config(state='normal')
            dlg.show_cb.config(state='normal')
            dlg.remember_cb.config(state='normal')
            dlg.entry.config(state='normal')
        except Exception:
            pass
        messagebox.showerror("Validation Failed", f"Could not validate the API key:\n\n{err_msg}")

    def on_success(saved_to_keyring):
        # Initialize real client
        init_client_with_key(key)
        # Write a success message into chat_history (do not print the key)
        chat_history.configure(state='normal')
        if remember_flag and saved_to_keyring:
            chat_history.insert(tk.END, "API key validated and saved to system keychain.\n\n")
        elif remember_flag and not saved_to_keyring:
            chat_history.insert(tk.END, "API key validated but could not be saved to keychain. Using it for this session only.\n\n")
        else:
            chat_history.insert(tk.END, "API key validated for this session (not saved to keychain).\n\n")
        chat_history.configure(state='disabled')
        chat_history.see(tk.END)

        messagebox.showinfo("Success", "API key validated successfully.")
        try:
            dlg.grab_release()
            dlg.destroy()
        except Exception:
            pass

        # Update UI widgets now that we have a valid key
        update_api_widgets()

    if err:
        app.after(0, on_fail)
        return

    # No error: optionally save to keyring
    saved = False
    if remember_flag:
        saved = save_api_key_in_keyring(key)

    app.after(0, lambda: on_success(saved))

def set_api_key_dialog(parent=None):
    dlg = tk.Toplevel(parent or app)
    dlg.title("Set API Key")
    dlg.grab_set()
    dlg.resizable(False, False)
    dlg.transient(parent or app)

    tk.Label(dlg, text="Enter your OpenAI API key:").grid(row=0, column=0, padx=12, pady=(12, 0), sticky='w')

    key_var = tk.StringVar()
    entry = tk.Entry(dlg, textvariable=key_var, width=60, show="*")
    entry.grid(row=1, column=0, padx=12, pady=6, sticky='we')

    # show/hide checkbox
    show_var = tk.BooleanVar(value=False)
    def toggle_show():
        entry.config(show="" if show_var.get() else "*")
    show_cb = tk.Checkbutton(dlg, text="Show key", variable=show_var, command=toggle_show)
    show_cb.grid(row=2, column=0, padx=12, sticky='w')

    # remember in keychain
    remember_var = tk.BooleanVar(value=True)
    remember_cb = tk.Checkbutton(dlg, text="Remember API key in system keychain", variable=remember_var)
    remember_cb.grid(row=3, column=0, padx=12, pady=(2, 6), sticky='w')

    # status label inside dialog
    status_var = tk.StringVar(value="")
    status_label = tk.Label(dlg, textvariable=status_var, anchor='w', fg='gray30')
    status_label.grid(row=4, column=0, padx=12, pady=(0, 6), sticky='we')

    btn_frame = tk.Frame(dlg)
    btn_frame.grid(row=5, column=0, padx=12, pady=(0, 12), sticky='e')

    def on_save():
        key = key_var.get().strip()
        if not key:
            messagebox.showwarning("No key", "Please enter a non-empty API key or click Cancel.")
            return
        # disable controls while validating
        status_var.set("Validating API key...")
        save_btn.config(state='disabled')
        cancel_btn.config(state='disabled')
        show_cb.config(state='disabled')
        remember_cb.config(state='disabled')
        entry.config(state='disabled')

        # Attach some references so worker can re-enable if needed
        dlg.status_var = status_var
        dlg.save_btn = save_btn
        dlg.cancel_btn = cancel_btn
        dlg.show_cb = show_cb
        dlg.remember_cb = remember_cb
        dlg.entry = entry

        t = threading.Thread(target=_validate_key_worker, args=(key, remember_var.get(), dlg), daemon=True)
        t.start()

    def on_cancel():
        try:
            dlg.grab_release()
            dlg.destroy()
        except Exception:
            pass

    save_btn = tk.Button(btn_frame, text="Validate & Save", command=on_save)
    save_btn.pack(side='right', padx=(6,0))
    cancel_btn = tk.Button(btn_frame, text="Cancel", command=on_cancel)
    cancel_btn.pack(side='right')

    # Pre-fill from any existing key (keyring or env)
    existing = load_api_key()
    if existing:
        key_var.set(existing)

    entry.focus_set()
    dlg.update_idletasks()
    dlg.minsize(dlg.winfo_reqwidth(), dlg.winfo_reqheight())

def forget_api_key():
    delete_api_key_from_keyring()
    # Do not forcibly clear session key (session-only key remains until restarted),
    # but re-init client without explicit key so SDK falls back to env var (if any).
    init_client_with_key(None)
    chat_history.configure(state='normal')
    chat_history.insert(tk.END, "API key removed from system keychain. If you set a key for this session it will remain active until the app restarts.\n\n")
    chat_history.configure(state='disabled')
    chat_history.see(tk.END)
    update_api_widgets()

def test_api_key():
    """
    Manual re-validation of the currently stored key (keyring or env or session).
    Runs in background and reports success/failure.
    """
    # Prefer session key, then keyring/env
    candidate = current_session_key or load_api_key()
    if not candidate:
        messagebox.showinfo("No API Key", "No API key is set. Use Key -> Set API Key first.")
        return

    def worker():
        try:
            temp_client = OpenAI(api_key=candidate)
            temp_client.chat.completions.create(
                model=_default_model,
                messages=[{"role": "user", "content": "Say ok in one word"}],
                max_completion_tokens=10
            )
        except Exception as exc:
            err_msg = str(exc)
            app.after(0, lambda msg=err_msg: messagebox.showerror("Test Failed", f"API key test failed:\n\n{msg}"))
            return
        app.after(0, lambda: messagebox.showinfo("Test Succeeded", "API key is valid."))

    threading.Thread(target=worker, daemon=True).start()

# -----------------------------
# Initialize the OpenAI client variable (we'll initialize properly later)
# -----------------------------
client = None

# Default model: prefer the config value if present and in the list, otherwise use first available
_default_model = DEFAULT_MODEL_VERSION if DEFAULT_MODEL_VERSION in AVAILABLE_MODELS else AVAILABLE_MODELS[0]

# -----------------------------
# Tkinter app state & helpers
# -----------------------------
conversation = [
    {"role": "system", "content": f"You are chatting with a user. Respond helpfully. Your name is {ASSISTANT_NAME}"}
]

conversation_lock = threading.Lock()

waiting = False
anim_after_id = None
anim_dots = 0

image_refs = []
default_image_folder = os.getcwd()

# === Image input (vision) state ===
# attached_images will hold dicts: {"orig_path":..., "bytes":..., "mime":...}
attached_images = []  # list of compressed image dicts selected to include with the next message

def model_supports_images(model_name):
    # Heuristic: support known vision models
    name = (model_name or "").lower()
    return (name in {m.lower() for m in VISION_CAPABLE_MODELS}) or ("4o" in name)

def compress_image_to_jpeg_bytes(path, max_dim=500, quality=75):
    """
    Open image at path, resize so max(width,height) <= max_dim, convert to RGB
    and save as JPEG into bytes. Returns (bytes, mime, was_resized). Raises on failure.
    """
    try:
        with Image.open(path) as img:
            w, h = img.size
            max_side = max(w, h)
            was_resized = max_side > max_dim
            if was_resized:
                ratio = max_dim / float(max_side)
                new_size = (int(w * ratio), int(h * ratio))
                img = img.resize(new_size, resample=RESAMPLE_FILTER)
            else:
                img = img.copy()

            # Convert to RGB to save as JPEG (flatten alpha if present)
            if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
                background = Image.new("RGB", img.size, (255, 255, 255))
                background.paste(img, mask=img.split()[-1])  # paste using alpha channel as mask
                img = background
            else:
                img = img.convert("RGB")

            buf = BytesIO()
            img.save(buf, format="JPEG", quality=quality, optimize=True)
            data = buf.getvalue()
            return data, "image/jpeg", was_resized
    except Exception:
        # Let caller handle failures (so attach_images can fallback and warn)
        raise

def image_bytes_to_data_url(img_bytes, mime):
    b64 = base64.b64encode(img_bytes).decode("ascii")
    return f"data:{mime};base64,{b64}"

def show_images_inline_from_bytes(byte_items):
    """
    Display images inline in the chat (used for user attachments preview and user message display).
    byte_items: iterable of dicts with keys {"bytes":..., "mime":..., "orig_path":...}
    """
    try:
        chat_history.configure(state='normal')
        for item in byte_items:
            try:
                b = item.get("bytes")
                if b is None:
                    # fallback: attempt to open orig_path
                    p = item.get("orig_path")
                    pil_img = Image.open(p)
                else:
                    pil_img = Image.open(BytesIO(b))

                chat_width = chat_history.winfo_width()
                if not chat_width or chat_width < 50:
                    max_display_width = 600
                else:
                    max_display_width = max(200, chat_width - 80)

                w, h = pil_img.size
                if w > max_display_width:
                    ratio = max_display_width / w
                    new_size = (int(w * ratio), int(h * ratio))
                    pil_img = pil_img.resize(new_size, resample=RESAMPLE_FILTER)

                tk_img = ImageTk.PhotoImage(pil_img)
                image_refs.append(tk_img)  # keep ref to avoid GC

                chat_history.image_create(tk.END, image=tk_img)
                chat_history.insert(tk.END, "\n")
            except Exception as e_img:
                chat_history.insert(tk.END, f"[Failed to preview image {os.path.basename(item.get('orig_path',''))}: {e_img}]\n")
        chat_history.insert(tk.END, "\n")
        chat_history.configure(state='disabled')
        chat_history.see(tk.END)
    except Exception as e_ui:
        chat_history.configure(state='normal')
        chat_history.insert(tk.END, f"Error displaying attached images: {e_ui}\n\n")
        chat_history.configure(state='disabled')
        chat_history.see(tk.END)

def attach_images():
    """
    Choose one or more image files to attach to the next user message.
    Compress/rescale them immediately and store the compressed bytes for preview and sending.
    If an image was larger than max_dim and compression failed, attach raw bytes but warn in chat.
    """
    global attached_images
    try:
        app.update()
        app.lift()
        app.focus_force()
    except Exception:
        pass

    paths = filedialog.askopenfilenames(
        parent=app,
        title="Attach image(s) for next message",
        initialdir=os.getcwd(),
        filetypes=[
            ("Image files", ("*.png","*.jpg","*.jpeg","*.webp","*.gif","*.bmp")),
            ("All files", "*.*"),
        ],
    )
    if not paths:
        return

    new_attached = []
    resized_count = 0
    unchanged_count = 0

    for p in paths:
        orig_max_side = None
        try:
            # attempt to read original dimensions (best-effort)
            with Image.open(p) as _img:
                ow, oh = _.size
                orig_max_side = max(ow, oh)
        except Exception:
            orig_max_side = None

        try:
            img_bytes, mime, was_resized = compress_image_to_jpeg_bytes(p, max_dim=MAX_UPLOADED_IMAGE_DIMENSION, quality=75)
            if was_resized:
                resized_count += 1
            else:
                unchanged_count += 1
            new_attached.append({"orig_path": p, "bytes": img_bytes, "mime": mime})
        except Exception as e:
            # compression/resizing failed — fallback to raw file bytes
            try:
                with open(p, "rb") as f:
                    raw = f.read()
                new_attached.append({"orig_path": p, "bytes": raw, "mime": "application/octet-stream"})
                unchanged_count += 1

                # If image looked larger than max_dim, warn that resize failed and an uncompressed file was attached
                if orig_max_side and orig_max_side > 500:
                    try:
                        size_bytes = os.path.getsize(p) if os.path.exists(p) else None
                        size_kb = f"{(size_bytes/1024):.1f} KB" if size_bytes else "unknown size"
                    except Exception:
                        size_kb = "unknown size"
                    chat_history.configure(state='normal')
                    chat_history.insert(
                        tk.END,
                        f"[Warning: could not resize '{os.path.basename(p)}' which was larger than 500px (attached uncompressed, {size_kb}). Sending it may incur high token usage.]\n",
                        "separator"
                    )
                    chat_history.configure(state='disabled')
                    chat_history.see(tk.END)
            except Exception as e2:
                chat_history.configure(state='normal')
                chat_history.insert(tk.END, f"[Failed to attach image {os.path.basename(p)}: {e2}]\n")
                chat_history.configure(state='disabled')
                chat_history.see(tk.END)

    if new_attached:
        attached_images = new_attached

        # Build an accurate summary message
        if resized_count == len(new_attached):
            summary = f"[Attached {len(attached_images)} image(s) for the next message (all resized to max 500px)]"
        elif resized_count == 0:
            summary = f"[Attached {len(attached_images)} image(s) for the next message (none needed resizing)]"
        else:
            summary = f"[Attached {len(attached_images)} image(s) for the next message ({resized_count} resized, {unchanged_count} unchanged; max 500px)]"

        chat_history.configure(state='normal')
        chat_history.insert(tk.END, summary + "\n", "separator")
        chat_history.configure(state='disabled')
        chat_history.see(tk.END)

        # Inline preview of the compressed images (or raw if fallback)
        show_images_inline_from_bytes(attached_images)

def clear_attached_images():
    """
    Clear any images previously attached for the next message.
    """
    global attached_images
    attached_images = []
    chat_history.configure(state='normal')
    chat_history.insert(tk.END, "[Cleared attached images]\n\n", "separator")
    chat_history.configure(state='disabled')
    chat_history.see(tk.END)

def animate_wait():
    global anim_dots, anim_after_id
    if not waiting:
        return
    anim_dots = (anim_dots + 1) % 4
    status_label.config(text="Thinking" + "." * anim_dots)
    anim_after_id = app.after(500, animate_wait)

def start_waiting():
    global waiting, anim_dots, anim_after_id
    if waiting:
        return
    waiting = True
    anim_dots = 0
    status_label.config(text="Thinking")
    anim_after_id = app.after(500, animate_wait)

def stop_waiting():
    global waiting, anim_after_id
    if not waiting:
        return
    waiting = False
    if anim_after_id is not None:
        try:
            app.after_cancel(anim_after_id)
        except Exception:
            pass
        anim_after_id = None
    status_label.config(text="Idle")

# -----------------------------
# Chat send (streaming) & image generation
# -----------------------------
def send_message():
    global attached_images
    user_message = user_input.get("1.0", tk.END).strip()

    # If neither text nor images, do nothing
    if not user_message and not attached_images:
        return

    selected_model = current_model_var.get()
    if attached_images and not model_supports_images(selected_model):
        # Soft warning in chat, but still allow sending (the API may reject it)
        chat_history.configure(state='normal')
        chat_history.insert(tk.END, f"[Warning: model '{selected_model}' may not support image input]\n\n", "separator")
        chat_history.configure(state='disabled')
        chat_history.see(tk.END)

    # Prepare the content for the API (text + image parts if any)
    content_parts = []
    if user_message:
        content_parts.append({"type": "text", "text": user_message})

    images_to_send = attached_images[:]  # copy then clear session attachments
    attached_images = []

    # Convert images (bytes) to data URLs and add as image parts
    for item in images_to_send:
        try:
            img_bytes = item.get("bytes")
            mime = item.get("mime", "image/jpeg")
            data_url = image_bytes_to_data_url(img_bytes, mime)
            content_parts.append({
                "type": "image_url",
                "image_url": {"url": data_url}
            })
        except Exception as e:
            # If an image fails to convert, mention it but continue
            chat_history.configure(state='normal')
            chat_history.insert(tk.END, f"[Failed to attach image {os.path.basename(item.get('orig_path',''))}: {e}]\n", "separator")
            chat_history.configure(state='disabled')
            chat_history.see(tk.END)

    # Disable the send button on the main thread immediately
    send_button.config(state='disabled')

    # Append to conversation
    with conversation_lock:
        if len(content_parts) == 1 and content_parts[0].get("type") == "text":
            conversation.append({"role": "user", "content": user_message})
        else:
            conversation.append({"role": "user", "content": content_parts})

    def show_user_message():
        chat_history.configure(state='normal')
        if user_message:
            chat_history.insert(tk.END, "You: " + user_message + "\n")
        else:
            chat_history.insert(tk.END, "You (images only):\n")
        chat_history.configure(state='disabled')
        chat_history.see(tk.END)
        # Show attached (compressed) images inline under the user's message
        if images_to_send:
            show_images_inline_from_bytes(images_to_send)
        user_input.delete("1.0", tk.END)

    app.after(0, show_user_message)
    start_waiting()

    def worker(model_to_use):
        try:
            with conversation_lock:
                messages_for_api = conversation.copy()

            stream = client.chat.completions.create(
                model=model_to_use,
                messages=messages_for_api,
                stream=True
            )

            def start_assistant_message():
                chat_history.configure(state='normal')
                sep = "\n--- Model response ---\n"
                chat_history.insert(tk.END, sep, "separator")
                chat_history.insert(tk.END, f"{ASSISTANT_NAME}: ")
                chat_history.configure(state='disabled')
                chat_history.see(tk.END)

            app.after(0, start_assistant_message)

            assistant_full = ""

            for event in stream:
                delta_text = ""
                try:
                    choice0 = event.choices[0]
                    delta = getattr(choice0, "delta", None)
                    if delta is None:
                        delta = choice0.get("delta", {})
                    if hasattr(delta, "content"):
                        delta_text = delta.content or ""
                    else:
                        delta_text = delta.get("content", "") if isinstance(delta, dict) else ""
                except Exception:
                    delta_text = ""

                if delta_text:
                    assistant_full += delta_text
                    def append_chunk(chunk=delta_text):
                        chat_history.configure(state='normal')
                        chat_history.insert(tk.END, chunk)
                        chat_history.configure(state='disabled')
                        chat_history.see(tk.END)
                    app.after(0, append_chunk)

            def finalize_assistant_message():
                chat_history.configure(state='normal')
                chat_history.insert(tk.END, "\n\n")
                chat_history.configure(state='disabled')
                chat_history.see(tk.END)

            app.after(0, finalize_assistant_message)

            with conversation_lock:
                conversation.append({"role": "assistant", "content": assistant_full})

        except Exception as e:
            print(e.__cause__)
            def show_error(err=e):  # bind e into default arg
                chat_history.configure(state='normal')
                chat_history.insert(tk.END, "Error: " + str(err) + "\n\n")
                chat_history.configure(state='disabled')
                chat_history.see(tk.END)

            app.after(0, show_error)

        finally:
            def on_request_complete():
                send_button.config(state='normal')
                stop_waiting()
            app.after(0, on_request_complete)

    t = threading.Thread(target=worker, args=(selected_model,), daemon=True)
    t.start()

# ------------------------------
# File operation functions
# ------------------------------

def export_chat():
    filename = f'chat_history-{datetime.now().strftime("%m%b%y-%H%M")}.txt'
    chat_content = chat_history.get('1.0', tk.END)
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(chat_content)
    print(f"Exported {filename}")

def read_text_file(path):
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        return f.read()

def read_docx_file(path):
    try:
        doc = Document(path)
    except Exception as e:
        raise RuntimeError(f"Error reading DOCX: {e}")
    paragraphs = [p.text for p in doc.paragraphs]
    return "\n".join(paragraphs)

def read_pdf_file(path):
    try:
        reader = PdfReader(path)
    except Exception as e:
        raise RuntimeError(f"Error opening PDF: {e}")
    text_chunks = []
    for page in reader.pages:
        try:
            text = page.extract_text()
        except Exception:
            text = None
        if text:
            text_chunks.append(text)
    return "\n\n".join(text_chunks)

def read_excel_file(path):
    try:
        import pandas as pd
        sheets = pd.read_excel(path, sheet_name=None)
        parts = []
        for sheet_name, df in sheets.items():
            parts.append(f"Sheet: {sheet_name}")
            parts.append(df.to_csv(sep='\t', index=False))
        return "\n\n".join(parts)
    except Exception as pandas_err:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"Sheet: {sheet_name}")
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join("" if v is None else str(v) for v in row)
                    rows.append(row_text)
                parts.append("\n".join(rows))
            return "\n\n".join(parts)
        except Exception as openpyxl_err:
            raise RuntimeError(
                "Failed to read Excel file. pandas error: "
                f"{pandas_err}; openpyxl error: {openpyxl_err}. "
                "To add Excel support install pandas and openpyxl (e.g. pip install pandas openpyxl)."
            )

def import_file():
    try:
        # Make sure main window is visible & focused for macOS
        try:
            app.update()
            app.lift()
            app.focus_force()
        except Exception:
            pass

        filepaths = filedialog.askopenfilenames(
            parent=app,
            title="Import file(s) for context",
            initialdir=os.getcwd(),
            filetypes=[
                ("All files", "*.*"),
                ("Text files", ("*.txt", "*.md")),
                ("Word documents", ("*.docx",)),
                ("PDF files", ("*.pdf",)),
                ("Excel files", ("*.xls", "*.xlsx"))
            ]
        )
        if not filepaths:
            return

        imported_count = 0
        preview_limit = 1000

        for filepath in filepaths:
            ext = os.path.splitext(filepath)[1].lower()
            try:
                if ext in ('.txt', '.md'):
                    content = read_text_file(filepath)
                elif ext == '.docx':
                    content = read_docx_file(filepath)
                elif ext == '.pdf':
                    content = read_pdf_file(filepath)
                elif ext in ('.xls', '.xlsx'):
                    content = read_excel_file(filepath)
                else:
                    # Try reading as text fallback
                    try:
                        content = read_text_file(filepath)
                    except Exception:
                        raise RuntimeError("Unsupported file type and not readable as text.")
            except Exception as e:
                chat_history.configure(state='normal')
                chat_history.insert(tk.END, f"Error importing file {os.path.basename(filepath)}: {e}\n\n")
                chat_history.configure(state='disabled')
                chat_history.see(tk.END)
                continue

            with conversation_lock:
                conversation.append({"role": "user", "content": content})

            preview = content if len(content) <= preview_limit else content[:preview_limit] + "\n...[truncated preview]"

            chat_history.configure(state='normal')
            chat_history.insert(tk.END, f"You (imported {os.path.basename(filepath)}):\n{preview}\n\n")
            chat_history.configure(state='disabled')
            chat_history.see(tk.END)

            imported_count += 1

        if imported_count == 0:
            messagebox.showinfo("Import Files", "No files were imported (all failed or canceled).")
        else:
            messagebox.showinfo("Import Files", f"Imported {imported_count} file(s).")
    except Exception as exc:
        print("import_file exception:", traceback.format_exc())
        messagebox.showerror("Import Error", f"Error opening file dialog or importing file:\n\n{exc}")

# ------------------------------
# Image generation functions
# ------------------------------
def generate_image_dialog():
    dlg = tk.Toplevel(app)
    dlg.title("Generate Image")
    dlg.grab_set()

    tk.Label(dlg, text="Enter image prompt:").grid(row=0, column=0, padx=8, pady=(8, 0), sticky='w')
    prompt_text = tk.Text(dlg, wrap=tk.WORD, width=60, height=6)
    prompt_text.grid(row=1, column=0, padx=8, pady=8, sticky='nsew')

    size_frame = tk.Frame(dlg)
    size_frame.grid(row=2, column=0, sticky='w', padx=8)
    tk.Label(size_frame, text="Size:").pack(side='left', padx=(0,6))
    size_var = tk.StringVar(value="1024x1024")
    size_choices = ["256x256", "512x512", "1024x1024"]
    size_menu = tk.OptionMenu(size_frame, size_var, *size_choices)
    size_menu.pack(side='left')

    folder_frame = tk.Frame(dlg)
    folder_frame.grid(row=3, column=0, sticky='we', padx=8, pady=(6, 0))
    tk.Label(folder_frame, text="Save folder:").pack(side='left')
    folder_label_var = tk.StringVar(value=default_image_folder)
    folder_label = tk.Label(folder_frame, textvariable=folder_label_var, anchor='w')
    folder_label.pack(side='left', padx=(6,8), fill='x', expand=True)

    def choose_folder():
        chosen = filedialog.askdirectory(title="Choose folder to save images", initialdir=folder_label_var.get() or os.getcwd())
        if chosen:
            folder_label_var.set(chosen)

    choose_btn = tk.Button(folder_frame, text="Choose...", command=choose_folder)
    choose_btn.pack(side='right')

    button_frame = tk.Frame(dlg)
    button_frame.grid(row=4, column=0, pady=(8, 8), sticky='e')

    def on_submit():
        prompt = prompt_text.get("1.0", tk.END).strip()
        if not prompt:
            return
        size = size_var.get()
        folder = folder_label_var.get() or os.getcwd()
        dlg.destroy()
        start_image_generation(prompt, size, folder)

    submit_btn = tk.Button(button_frame, text="Generate", command=on_submit)
    submit_btn.pack(side='right', padx=5)
    cancel_btn = tk.Button(button_frame, text="Cancel", command=dlg.destroy)
    cancel_btn.pack(side='right')

    dlg.update_idletasks()
    dlg.minsize(dlg.winfo_reqwidth(), dlg.winfo_reqheight())

def start_image_generation(prompt, size, folder):
    with conversation_lock:
        conversation.append({"role": "user", "content": f"[Image prompt] {prompt}"})

    app.after(0, lambda: show_user_prompt_ui(prompt, size))
    start_waiting()
    send_button.config(state='disabled')

    t = threading.Thread(target=image_worker, args=(prompt, size, folder), daemon=True)
    t.start()

def image_worker(prompt, size, folder):
    try:
        os.makedirs(folder, exist_ok=True)

        response = client.images.generate(
            model=IMAGE_MODEL_VERSION,
            prompt=prompt,
            size=size
        )

        img_b64 = None
        img_url = None
        try:
            first = response.data[0]
            img_b64 = getattr(first, "b64_json", None) or (first.get("b64_json") if isinstance(first, dict) else None)
            img_url = getattr(first, "url", None) or (first.get("url") if isinstance(first, dict) else None)
        except Exception:
            pass

        saved_filename = None
        if img_b64:
            try:
                img_bytes = base64.b64decode(img_b64)
                saved_filename = os.path.join(folder, f"image-{datetime.now().strftime('%Y%m%d-%H%M%S')}.png")
                with open(saved_filename, "wb") as f:
                    f.write(img_bytes)
            except Exception as e:
                raise RuntimeError(f"Failed to decode/save image: {e}")
        elif img_url:
            try:
                saved_filename = os.path.join(folder, f"image-{datetime.now().strftime('%Y%m%d-%H%M%S')}.png")
                urllib.request.urlretrieve(img_url, saved_filename)
            except Exception as e:
                raise RuntimeError(f"Failed to download image from URL: {e}")
        else:
            raise RuntimeError(f"No image data returned. Response: {response}")

        assistant_text = f"[Image generated: {saved_filename}] (prompt: {prompt})"
        with conversation_lock:
            conversation.append({"role": "assistant", "content": assistant_text})

        app.after(0, lambda: show_assistant_image_ui(saved_filename, assistant_text))

    except Exception as e:
        print("Image generation error:", traceback.format_exc())
        app.after(0, lambda: show_image_error_ui(e))

    finally:
        app.after(0, cleanup_after_image)

def show_user_prompt_ui(prompt, size):
    chat_history.configure(state='normal')
    chat_history.insert(tk.END, f"You (image prompt, {size}): {prompt}\n")
    chat_history.configure(state='disabled')
    chat_history.see(tk.END)

def show_assistant_image_ui(saved_filename, assistant_text):
    try:
        chat_history.configure(state='normal')
        sep = "\n--- Model response ---\n"
        chat_history.insert(tk.END, sep, "separator")
        chat_history.insert(tk.END, f"{ASSISTANT_NAME}: {assistant_text}\n")

        pil_img = Image.open(saved_filename)

        chat_width = chat_history.winfo_width()
        if not chat_width or chat_width < 50:
            max_display_width = 600
        else:
            max_display_width = max(200, chat_width - 80)

        w, h = pil_img.size
        if w > max_display_width:
            ratio = max_display_width / w
            new_size = (int(w * ratio), int(h * ratio))
            pil_img = pil_img.resize(new_size, resample=RESAMPLE_FILTER)

        tk_img = ImageTk.PhotoImage(pil_img)
        image_refs.append(tk_img)

        chat_history.image_create(tk.END, image=tk_img)
        chat_history.insert(tk.END, "\n\n")
        chat_history.configure(state='disabled')
        chat_history.see(tk.END)

    except Exception as e_ui:
        chat_history.configure(state='normal')
        chat_history.insert(tk.END, f"Error displaying image: {e_ui}\n\n")
        chat_history.configure(state='disabled')
        chat_history.see(tk.END)

def show_image_error_ui(exc):
    chat_history.configure(state='normal')
    chat_history.insert(tk.END, "Error generating image: " + str(exc) + "\n\n")
    chat_history.configure(state='disabled')
    chat_history.see(tk.END)

def cleanup_after_image():
    send_button.config(state='normal')
    stop_waiting()

# ------------------------------
# Model menu change handler
# ------------------------------
def on_model_change():
    selected = current_model_var.get()
    chat_history.configure(state='normal')
    chat_history.insert(tk.END, f"[Model switched to: {selected}]\n\n")
    chat_history.configure(state='disabled')
    chat_history.see(tk.END)

# ------------------------------
# Build the GUI
# ------------------------------
app = tk.Tk()
app.title("GPT GUIde")

# Default UI model selection variable
current_model_var = tk.StringVar(app, value=_default_model)

menu = Menu(app)
app.config(menu=menu)

# File menu
filemenu = Menu(menu, tearoff=0)
menu.add_cascade(label="File", menu=filemenu)
filemenu.add_command(label="Export Chat", command=export_chat)
filemenu.add_command(label="Import File(s)", command=import_file)
filemenu.add_command(label="Generate Image...", command=generate_image_dialog)
# New: attach/clear image inputs for chat
filemenu.add_command(label="Attach Image(s)...", command=attach_images)
filemenu.add_command(label="Clear Attached Images", command=clear_attached_images)
# capture index of generate image command so we can enable/disable it later
_generate_image_menu_index = filemenu.index("end") - 2  # adjust since two items were added after

# Model menu
modelmenu = Menu(menu, tearoff=0)
menu.add_cascade(label="Model", menu=modelmenu)
for m in AVAILABLE_MODELS:
    modelmenu.add_radiobutton(label=m, variable=current_model_var, value=m, command=on_model_change)

# Key menu (Set / Forget / Test)
keymenu = Menu(menu, tearoff=0)
menu.add_cascade(label="Key", menu=keymenu)
keymenu.add_command(label="Set API Key", command=lambda: set_api_key_dialog(app))
keymenu.add_command(label="Forget API Key", command=forget_api_key)
keymenu.add_command(label="Test API Key", command=test_api_key)

# Configure grid weights
app.grid_rowconfigure(0, weight=1)
app.grid_rowconfigure(1, weight=0)
app.grid_rowconfigure(2, weight=0)
app.grid_rowconfigure(3, weight=0)
app.grid_columnconfigure(0, weight=1)
app.grid_columnconfigure(1, weight=0, minsize=80)

chat_history = scrolledtext.ScrolledText(app, state='disabled', wrap=tk.WORD)
chat_history.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky='nsew')

default_font = tkfont.nametofont("TkDefaultFont")
sep_font = default_font.copy()
try:
    sep_font.configure(size=max(default_font.cget("size") - 2, 8), slant='italic')
except Exception:
    sep_font = ("TkDefaultFont", max(default_font.cget("size") - 2, 8), "italic")
chat_history.tag_configure("separator", font=sep_font, foreground="gray")

user_input = tk.Text(app, wrap=tk.WORD, height=4)
user_input.grid(row=1, column=0, padx=10, pady=10, sticky='nsew')

send_button = tk.Button(app, text="Send", command=send_message)
send_button.grid(row=1, column=1, padx=10, pady=10, sticky='ns')

status_label = tk.Label(app, text="Idle", anchor='w')
status_label.grid(row=2, column=0, columnspan=2, sticky='ew', padx=10, pady=(0,2))

credit_font = default_font.copy()
try:
    credit_font.configure(size=max(default_font.cget("size") - 3, 8), slant='italic')
except Exception:
    credit_font = ("TkDefaultFont", max(default_font.cget("size") - 3, 8), "italic")
credit_label = tk.Label(app, text="Created by David Miles", anchor='e', font=credit_font, foreground="gray30")
credit_label.grid(row=3, column=0, columnspan=2, sticky='ew', padx=10, pady=(0,10))

app.update_idletasks()
min_w = app.winfo_reqwidth()
min_h = app.winfo_reqheight()
app.minsize(min_w, min_h)
app.geometry(f"{min_w}x{min_h}")

# ------------------------------
# API key initialization + widget enabling
# ------------------------------
# Initialize client using any existing key (keyring or env)
initial_key = load_api_key()
init_client_with_key(initial_key)

def update_api_widgets():
    """
    Enable or disable UI elements that require an API key.
    Criteria: session key present (current_session_key) OR environment variable present.
    """
    has_env_key = bool(os.getenv("OPENAI_KEY"))
    has_key_for_session = bool(current_session_key) or has_env_key

    # Send button
    try:
        send_button.config(state='normal' if has_key_for_session else 'disabled')
    except Exception:
        pass

    # Generate Image menu entry (in File menu), using the captured index
    try:
        filemenu.entryconfig(_generate_image_menu_index, state='normal' if has_key_for_session else 'disabled')
    except Exception:
        pass

# Warn in chat if no key present and update widgets accordingly
def check_api_key_and_warn():
    if load_api_key() is None and not current_session_key and not os.getenv("OPENAI_KEY"):
        chat_history.configure(state='normal')
        chat_history.insert(tk.END, "Warning: No API key set. OpenAI requests will fail until you set a key under Key -> Set API Key.\n\n")
        chat_history.configure(state='disabled')
        chat_history.see(tk.END)
    update_api_widgets()

# Call initial check now that widgets exist
check_api_key_and_warn()

# Start main loop
app.mainloop()